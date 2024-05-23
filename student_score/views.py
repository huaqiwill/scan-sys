import json
import re
from collections import OrderedDict
from itertools import chain, zip_longest
from openpyxl import load_workbook
import numpy as np
import pandas as pd

from django.core.paginator import Paginator
from django.db.models import Count, Sum, Avg, Q, Max
from django.http import JsonResponse
from django.shortcuts import render, redirect, HttpResponse
from pandas.core.reshape import pivot
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response

from common.API import res_josn_data
from common.API.auth import login_required, authorize
from student_score import models
from student_score.utils.bootstrap import BootStrapModelForm
from rest_framework import views
from student_score.management import serializer as ser

from student_score import models as m_model
from student_score.models import StudentCourseMemberShip


# Create your views here.


class Pagination(PageNumberPagination):
    '''
    自定义分页
    '''

    # 默认每页显示的个数
    page_size = 20
    # 可以动态改变每页显示的个数
    page_size_query_param = 'limit'
    # 页码参数
    page_query_param = 'page'
    # 最多能显示多少页
    max_page_size = 100

    def get_paginated_response(self, data):
        return Response(OrderedDict([
            ('count', self.page.paginator.count),
            ('next', self.get_next_link()),
            ('previous', self.get_previous_link()),
            ('code', 0),
            ('msg', '数据获取成功！'),
            ('data', data),
        ]))


def college_manage(request):
    return render(request, 'student_score/college_main.html')


def college_add(request):
    return render(request, 'student_score/college_add.html')


class CollegeManage(views.APIView):
    """
    学院管理
    """

    def get(self, request):
        """
        获取学院列表
        """
        try:
            school = request.session.get("department")
            role_des = request.session.get("role_des")

            if role_des != "管理员":
                datas = m_model.College.objects.filter(name=school)
            else:
                datas = m_model.College.objects.all()
            obj = ser.CollegeSerializer(datas, many=True)
            return Response({
                'code': 0,
                'data': obj.data,
                'message': '获取学院列表数据成功！'
            })

        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'message': '获取学院列表数据失败！'
            })

    def post(self, request):
        """
        新增学院信息
        """
        try:
            role_des = request.session.get("role_des")
            name = request.data.get('name')
            if role_des == "管理员":
                if not name:
                    return Response({
                        'code': 500,
                        'data': None,
                        'msg': '学院名字不能为空！'
                    })
                colleges = m_model.College.objects.filter(name=name)
                if colleges:
                    return Response({
                        'code': 500,
                        'data': None,
                        'msg': '学院已存在！'
                    })
                m_model.College.objects.create(name=name)
                return Response({
                    'code': 0,
                    'data': None,
                    'msg': f'学校:{name} 新增成功！'
                })
            else:
                return Response({
                    'code': 0,
                    'data': None,
                    'msg': '你没有权限增加学校！'
                })

        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '新增学院失败！'
            })

    def delete(self, request):
        """
        删除学校信息
        """
        try:
            user_list = []
            post_data_str = request.POST.get('Params', None)
            post_data = json.loads(post_data_str)
            if not post_data:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '学校不存在！'
                })
            if isinstance(post_data, dict):  # 判断单条删除（字典数据）
                db_id = post_data["id"]
                m_model.College.objects.filter(id=db_id).delete()
            else:
                for item in post_data:  # 批量删除（列表数据）
                    db_id = item['id']
                    user_name = item['name']
                    m_model.College.objects.filter(id=db_id).delete()
                    user_list.append(user_name)

            return Response({
                'code': 0,
                'data': None,
                'msg': '删除学校成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '删除学校失败！'
            })

    def put(self, request):
        """
        修改学院信息
        """
        try:
            college_id = request.data.get('dbID')
            college_name = request.data.get('value')
            print(request.data)
            print(college_id)
            if not college_id:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '学院id不能为空！'
                })
            college = m_model.College.objects.get(id=college_id)
            college.name = college_name
            college.save()
            return Response({
                'code': 200,
                'data': None,
                'msg': '修改学院成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '修改学院失败！'
            })


def grade_manage(request):
    return render(request, 'student_score/grade_main.html')


def grade_add(request):
    return render(request, 'student_score/grade_add.html')


class GradeManage(views.APIView):
    """
    年级管理
    """

    def get(self, request):
        """
        获取年级列表
        """
        try:
            limit = request.GET.get("limit")  # 每页数量
            page_num = request.GET.get("page")  # 第几页
            school = request.session.get("department")
            role_des = request.session.get("role_des")
            college = m_model.College.objects.filter(name=school).first()
            if role_des != "管理员":
                # colleges = m_model.College.objects.filter(id=college_id)
                # majors = m_model.Major.objects.filter(college=colleges[0], name=major_name)
                grades = m_model.Grade.objects.filter(college_id=college.id)
            else:
                grades = m_model.Grade.objects.all()

            datas = ser.GradeSerializer(grades, many=True)
            print(grades)
            return Response({
                'code': 0,
                'data': datas.data,
                'msg': '获取年级数据成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '获取年级数据失败！'
            })

    def post(self, request):
        """
        新增或者修改年级信息
        """

        try:
            school = request.session.get("department")
            college = m_model.College.objects.filter(name=school).first()
            collegeID = college.id
            gradeName = request.data.get('grade')
            print(collegeID, gradeName)
            if not collegeID:
                return Response({
                    'code': 0,
                    'data': None,
                    'msg': '学校不能为空！'
                })
            if not gradeName:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '年级不能为空！'
                })

            colleges = m_model.College.objects.filter(id=collegeID)
            # majors = m_model.Major.objects.filter(name=majorName, college=colleges[0])
            grades = m_model.Grade.objects.filter(name=gradeName, college=colleges[0])
            if grades:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '年级已存在！'
                })
            m_model.Grade.objects.create(name=gradeName, college=colleges[0])
            return Response({
                'code': 0,
                'data': None,
                'msg': '新增年级成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '新增年级失败！'
            })

    def delete(self, request):
        """
        删除班级信息
        """
        try:
            user_list = []
            post_data_str = request.POST.get('Params', None)
            post_data = json.loads(post_data_str)
            if not post_data:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '班级不存在！'
                })
            if isinstance(post_data, dict):  # 判断单条删除（字典数据）
                db_id = post_data["id"]
                m_model.Grade.objects.filter(id=db_id).delete()
            else:
                for item in post_data:  # 批量删除（列表数据）
                    db_id = item['id']
                    user_name = item['name']
                    m_model.Grade.objects.filter(id=db_id).delete()
                    user_list.append(user_name)

            return Response({
                'code': 0,
                'data': None,
                'msg': '删除班级成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '删除班级失败！'
            })

    def put(self, request):
        """
        新增或者修改年级信息
        """
        try:
            gradeName = request.data.get('gradeName')
            gradeId = request.data.get('gradeId')
            grades = m_model.Grade.objects.filter(id=gradeId)
            if not grades:
                return Response({
                    'code': 500,
                    'data': None,
                    'message': '年级不存在！'
                })
            grades.update(name=gradeName)
            return Response({
                'code': 200,
                'data': None,
                'message': '修改年级成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'message': '修改年级失败！'
            })


def class_manage(request):
    return render(request, 'student_score/class_main.html')


def class_add(request):
    return render(request, 'student_score/class_add.html')


class ClassManage(views.APIView):
    """
    班级管理
    """

    def get(self, request):
        """
        获取班级列表
        """
        try:
            grade_id = request.query_params.get('grade_id')  # 获取新增学生页面，年级下拉菜单改变时过来的gradeid

            # major_name = request.query_params.get('majorName')
            school = request.session.get("department")
            role_des = request.session.get("role_des")
            college = m_model.College.objects.filter(name=school).first()
            if grade_id:
                grades = m_model.Class.objects.filter(grade_id=grade_id)
                datas = ser.ClassSerializer(grades, many=True)
                return Response({
                    'code': 0,
                    'data': datas.data,
                    'msg': '获取班级数据成功！'
                })
            if role_des != "管理员":
                grades = m_model.Class.objects.filter(grade__college_id=college.id)
                # majors = m_model.Major.objects.filter(college=colleges[0], name=major_name)
                # grades = m_model.Class.objects.filter(grade=grades[0])
            else:
                grades = m_model.Class.objects.all()
                print(grades)
            datas = ser.ClassSerializer(grades, many=True)

            return Response({
                'code': 0,
                'data': datas.data,
                'msg': '获取班级数据成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '获取年级数据失败！'
            })

    def post(self, request):
        """
        新增或者修改班级信息
        """

        try:

            school = request.session.get("department")
            college = m_model.College.objects.filter(name=school).first()
            college_id = college.id
            grade_id = request.data.get('grade')
            className = request.data.get('class')

            if not college_id:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '学校不能为空！'
                })
            if not grade_id:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '年级不能为空！'
                })

            colleges = m_model.College.objects.filter(id=college_id)
            grades = m_model.Grade.objects.filter(id=grade_id, college=colleges[0])
            t_class = m_model.Class.objects.filter(name=className, grade=grades[0])
            if t_class:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '班级已存在！'
                })
            m_model.Class.objects.create(name=className, grade_id=grade_id)
            return Response({
                'code': 0,
                'data': None,
                'msg': '新增班级成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '新增班级失败！'
            })

    def delete(self, request):
        """
        删除班级信息
        """
        try:
            user_list = []
            post_data_str = request.POST.get('Params', None)
            post_data = json.loads(post_data_str)
            if not post_data:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '班级不存在！'
                })
            if isinstance(post_data, dict):  # 判断单条删除（字典数据）
                db_id = post_data["id"]
                m_model.Class.objects.filter(id=db_id).delete()
            else:
                for item in post_data:  # 批量删除（列表数据）
                    db_id = item['id']
                    user_name = item['name']
                    m_model.Class.objects.filter(id=db_id).delete()
                    user_list.append(user_name)

            return Response({
                'code': 0,
                'data': None,
                'msg': '删除班级成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '删除班级失败！'
            })

    def put(self, request):
        """
        新增或者修改年级信息
        """
        try:
            gradeName = request.data.get('gradeName')
            gradeId = request.data.get('gradeId')
            grades = m_model.Grade.objects.filter(id=gradeId)
            if not grades:
                return Response({
                    'code': 500,
                    'data': None,
                    'message': '年级不存在！'
                })
            grades.update(name=gradeName)
            return Response({
                'code': 200,
                'data': None,
                'message': '修改年级成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'message': '修改年级失败！'
            })


def course_manage(request):
    return render(request, 'student_score/course_main.html')


def course_add(request):
    return render(request, 'student_score/course_add.html')


class courseManage(views.APIView):
    """
    课程管理
    """

    def get(self, request):
        """
        获取班级列表
        """
        try:
            school = request.session.get("department")
            role_des = request.session.get("role_des")
            college = m_model.College.objects.filter(name=school).first()
            # 当前页
            page = request.GET.get('page', 1)
            # 每页条数
            size = request.GET.get('limit', 1)
            # 开始位置
            data_start = (int(page) - 1) * int(size)
            # 结束位置
            data_end = (int(page)) * int(size)
            if role_des != "管理员":
                grades = m_model.Course.objects.filter(grade__college_id=college.id)[data_start:data_end].values('id',
                                                                                                                 'grade__college__name',
                                                                                                                 'grade__name',
                                                                                                                 'name')
            else:
                grades = m_model.Course.objects.all()[data_start:data_end].values('id', 'grade__college__name',
                                                                                  'grade__name', 'name')
            # datas = ser.CourseSerializer(grades, many=True)
            count = m_model.Course.objects.count()
            return Response({
                'code': 0,
                'data': grades,
                'count': count,
                'msg': '获取班级数据成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '获取年级数据失败！'
            })

    def post(self, request):
        """
        新增或者修改课程信息
        """

        try:
            school = request.session.get("department")
            college = m_model.College.objects.filter(name=school).first()
            college_id = college.id
            grade_id = request.data.get('grade')
            courseName = request.data.get('course')

            if not college_id:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '学校不能为空！'
                })
            if not grade_id:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '年级不能为空！'
                })

            colleges = m_model.College.objects.filter(id=college_id)
            grades = m_model.Grade.objects.filter(id=grade_id, college=colleges[0])
            courses = m_model.Course.objects.filter(name=courseName, grade=grades[0])

            if courses:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '此课程已存在！'
                })
            course = m_model.Course.objects.create(name=courseName, grade_id=grade_id)

            # 这个专业-年级-所有的班级下面的学生，自动绑定和这个课程的关系

            classes = grades[0].t_class.all()
            print(classes)
            for t_class in classes:
                students = t_class.class_student.all()
                print(students)
                for stu in students:
                    print(stu.id, course.id, college_id)
                    # obj=models.StudentCourseMemberShip.objects.create(student_id=stu.id, course=course, course_id=college_id)
                    m_model.StudentCourseMemberShip.objects.create(course=course, student=stu, college_id=college_id)

            return Response({
                'code': 0,
                'data': None,
                'msg': '新增课程成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '新增课程失败！'
            })

    def delete(self, request):
        """
        删除课程信息
        """
        try:
            # user_list = []
            post_data_str = request.POST.get('Params', None)
            post_data = json.loads(post_data_str)
            print(post_data)
            if not post_data:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '课程不存在！'
                })
            if isinstance(post_data, dict):  # 判断单条删除（字典数据）
                db_id = post_data["id"]
                m_model.Course.objects.filter(id=db_id).delete()
            else:
                for item in post_data:  # 批量删除（列表数据）
                    db_id = item['id']
                    # user_name = item['name']
                    m_model.Course.objects.filter(id=db_id).delete()
                    # user_list.append(user_name)

            return Response({
                'code': 0,
                'data': None,
                'msg': '删除课程成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '删除课程失败！'
            })

    def put(self, request):
        """
        新增或者修改年级信息
        """
        try:
            gradeName = request.data.get('gradeName')
            gradeId = request.data.get('gradeId')
            grades = m_model.Grade.objects.filter(id=gradeId)
            if not grades:
                return Response({
                    'code': 500,
                    'data': None,
                    'message': '年级不存在！'
                })
            grades.update(name=gradeName)
            return Response({
                'code': 200,
                'data': None,
                'message': '修改年级成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'message': '修改年级失败！'
            })


def student_manage(request):
    return render(request, 'student_score/student_main.html')


def student_add(request):
    return render(request, 'student_score/student_add.html')


def student_upload_add(request):
    return render(request, 'student_score/student_upload_add.html')


class studentManage(views.APIView):
    """
    学生管理
    """

    def get(self, request):
        """
        获取学生列表
        """
        try:
            # college_name = request.query_params.get('collegeName')
            # # major_name = request.query_params.get('majorName')
            school = request.session.get("department")
            college = m_model.College.objects.filter(name=school).first()
            if college:
                #     # colleges = m_model.College.objects.filter(name=school)
                #     # majors = m_model.Major.objects.filter(college=colleges[0], name=major_name)
                grades = m_model.Student.objects.filter(t_class__grade__college_id=college.id)
            #
            else:
                grades = m_model.Student.objects.all()

            # 数据分页
            pg = Pagination()
            pager_roles = pg.paginate_queryset(queryset=grades, request=request, view=self)
            datas = ser.StudentSerializer(instance=pager_roles, many=True)
            #  用get_paginated_response方法，前端拿到的返回值中有数据总条数，上一页的链接，下一页的链接。
            return pg.get_paginated_response(datas.data)
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '获取学生数据失败！'
            })

    def post(self, request):
        """
        新增或者修改年级信息
        """

        try:
            school = request.session.get("department")
            college = m_model.College.objects.filter(name=school).first()
            college_id = college.id
            grade_id = request.data.get('grade')
            class_id = request.data.get('class')
            Name = request.data.get('name')
            # print(college_id,grade_id,className,Name)
            if not college_id:
                return Response({
                    'code': 0,
                    'data': None,
                    'msg': '学校不能为空！'
                })
            if not grade_id:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '年级不能为空！'
                })
            if not class_id:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '班级不能为空！'
                })
            if not Name:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '名字不能为空！'
                })
            colleges = m_model.College.objects.filter(id=college_id)
            grades = m_model.Grade.objects.filter(id=grade_id, college=colleges[0])
            classs = m_model.Class.objects.filter(id=class_id, grade=grades[0])
            names = m_model.Student.objects.filter(name=Name, t_class=classs[0])
            if names:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '该学生已存在！'
                })
            stu = m_model.Student.objects.create(name=Name, t_class=classs[0])
            print(stu.pk)
            # 新建数据到成绩表StudentCourseMemberShip 建立并关联所有课程的学生信息，自动绑定年级-课程（多个课程）-学生的关系
            print(grade_id)

            course = m_model.Course.objects.filter(grade_id=grade_id)  # 年级获取课程

            for itm in course:
                models.StudentCourseMemberShip.objects.create(student_id=stu.id, course_id=itm.id,
                                                              college_id=college_id)
            return Response({
                'code': 0,
                'data': None,
                'msg': '新增学生成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '新增学生失败！'
            })

    def delete(self, request):
        """
        删除学生信息
        """
        try:
            # user_list = []
            post_data_str = request.POST.get('Params', None)
            post_data = json.loads(post_data_str)
            if not post_data:
                return Response({
                    'code': 500,
                    'data': None,
                    'msg': '学生不存在！'
                })
            if isinstance(post_data, dict):  # 判断单条删除（字典数据）
                db_id = post_data["id"]
                m_model.Student.objects.filter(id=db_id).delete()
            else:
                for item in post_data:  # 批量删除（列表数据）
                    db_id = item['id']
                    # user_name = item['name']
                    m_model.Student.objects.filter(id=db_id).delete()
                    # user_list.append(user_name)

            return Response({
                'code': 0,
                'data': None,
                'msg': '删除学生成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '删除学生失败！'
            })

    def put(self, request):
        """
        新增或者修改年级信息
        """
        try:
            gradeName = request.data.get('gradeName')
            gradeId = request.data.get('gradeId')
            grades = m_model.Grade.objects.filter(id=gradeId)
            if not grades:
                return Response({
                    'code': 500,
                    'data': None,
                    'message': '年级不存在！'
                })
            grades.update(name=gradeName)
            return Response({
                'code': 200,
                'data': None,
                'message': '修改年级成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'message': '修改年级失败！'
            })


def student_multi(request):
    """ 批量导入（Excel文件）"""
    global college_id

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("file")
    print(file_object)

    if file_object:
        # 2.对象传递给openpyxl，由openpyxl读取文件的内容
        wb = load_workbook(file_object)
        sheet = wb.worksheets[0]
        chongfu_obj_list = []
        fenshu_obj_list = []
        lit = []
        i = 0
        school = request.session.get('department')
        print((school))

        # 3.循环获取每一行数据
        for row in sheet.iter_rows(min_row=2):
            college = school
            grade = row[0].value
            t_class = row[1].value
            name = row[2].value
            print(grade)
            # 查询是否存在重复数据,如有一条数据存在重复则全部不导入
            colleges = models.College.objects.filter(name=college)

            if colleges:
                grades = models.Grade.objects.filter(name=grade, college=colleges[0])
                if grades:
                    classes = models.Class.objects.filter(name=t_class, grade=grades[0])
                    if classes:
                        student = models.Student.objects.filter(name=name, t_class=classes[0]).first()
                        if student:
                            i += 1
                            chongfu_obj_list.append(student.name)
                        else:
                            stu_list = {}
                            stu_list['name'] = name
                            stu_list['colleges'] = colleges[0]
                            stu_list['grades'] = grades[0]
                            stu_list['classes'] = classes[0]
                            lit.append(stu_list)
                    else:
                        return JsonResponse(
                            {'code': 0, 'msg': "导入的班级在系统中不存在，请新建班级或重新修改模板后再上传！"})
                else:
                    return JsonResponse({'code': 0, 'msg': "导入的年级在系统中不存在，请新建年级或重新修改模板后上传！"})
            else:
                JsonResponse({'code': 0, 'msg': "导入的学校在系统中不存在，请新建学校或重新修改模板后上传！"})

        if i > 0:
            return JsonResponse({'code': 0, 'msg': '以下学生存在重复：{}'.format(chongfu_obj_list)},
                                json_dumps_params={'ensure_ascii': False})

        for itm in lit:
            course = models.Course.objects.filter(grade=itm['grades'])  # 年级获取课程
            if course:
                stu = models.Student.objects.create(name=itm['name'], t_class=itm['classes'])
                # 新建数据到成绩表StudentCourseMemberShip 建立并关联所有课程的学生信息，自动绑定年级-课程（多个课程）-学生的关系

                for i in course:
                    fenshu = StudentCourseMemberShip(
                        student_id=stu.id,
                        course_id=i.id,
                        college_id=colleges.first().id
                    )
                    fenshu_obj_list.append(fenshu)
            else:

                return JsonResponse({'code': 0, 'msg': "此年级没有建立课程: {}".format(itm['grades'])})
        # 批量保存数据
        print(fenshu_obj_list)
        models.StudentCourseMemberShip.objects.bulk_create(fenshu_obj_list)
        response_data = {'code': 0, 'msg': "上传成功！"}
        return JsonResponse(response_data)
    return JsonResponse({'code': 0, 'msg': '上传失败！'})


def score_manage(request):
    # 把查询条件传到前端
    queryset = m_model.College.objects.all()
    grade = request.session.get('position')  # 用户录入分工——年级
    course = request.session.get('email')  # 用户录入分工——学科
    context = {
        "queryset": queryset,
        "grade": grade,
        "course": course,
    }
    return render(request, 'student_score/score_main.html', context)


class ScoreManage(views.APIView):
    """
    成绩管理
    """

    def get(self, request):
        """
        获取成绩列表
        """
        try:
            # data_str = request.POST.get('Params', None)
            # post_data = json.loads(data_str)
            # user_status = post_data['status']
            # user_role = post_data['role']
            # print(user_status)

            school = request.session.get("department")
            role_des = request.session.get("role_des")
            college = m_model.College.objects.filter(name=school).first()
            post_data_str = request.POST.get('Params', None)
            print(post_data_str)
            if post_data_str is None:
                scores = m_model.StudentCourseMemberShip.objects.all()

            # 当前页
            page = request.GET.get('page', 1)
            # 每页条数
            size = request.GET.get('limit', 1)
            # 开始位置
            data_start = (int(page) - 1) * int(size)
            # 结束位置
            data_end = (int(page)) * int(size)
            tea = StudentCourseMemberShip.objects.all()[data_start:data_end].values('college__name',
                                                                                    'college__grade__name',
                                                                                    'student__t_class__name',
                                                                                    'course__name', 'student__name',
                                                                                    'result')
            print(tea)
            # 获取总条数
            count = m_model.StudentCourseMemberShip.objects.count()
            # datas = ser.ScoreSerializer(instance=tea, many=True)
            # ser = ShowStuSer(tea, many=True)
            return Response({
                'code': 0,
                'data': tea,
                'count': count
            })

            # else:
            #     post_data = json.loads(post_data_str)
            #     print(post_data)
            #     role_value = post_data['roleID']
            #     code_name = post_data['codeName']
            #     role_name = post_data['roleName']
            #     role_enable = post_data['enable']
            #     role_remark = post_data['remark']
            #
            #     filters = {}  # 查询参数构造
            #     # model或数据库对应字段
            #     orm_field = ['__gt', '__gte', '__lt', '__lte', '__exact', '__iexact', '__contains', '__icontains',
            #                  '__startswith', '__istartswith', '__endswith', '__iendswith', '__range', '__isnull',
            #                  '__in']
            #     filed_dict = {0: 'role_value', 1: 'name', 2: 'code', 3: 'enable', 4: 'remark'}
            #     param_list = [role_value, role_name, code_name, role_enable, role_remark]
            #
            #     for i in range(len(param_list)):
            #         if param_list[i] not in (None, ''):
            #             db_field = filed_dict[i] + orm_field[7]
            #             filters[db_field] = param_list[i]
            #
            #     print('filters:', filters)
            #
            #     user_obj = Role.objects.filter(**filters).order_by('id')
            # # 数据分页
            # pg = Pagination()
            # print(pg)
            # pager_roles = pg.paginate_queryset(queryset=scores, request=request, view=self)
            # print(pager_roles)
            # datas = ser.ScoreSerializer(instance=tea, many=True)
            # #  用get_paginated_response方法，前端拿到的返回值中有数据总条数，上一页的链接，下一页的链接。
            # return pg.get_paginated_response(datas.data)

        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'message': '获取学生课程成绩数据失败！'
            })

    def post(self, request):
        """新增数据"""
        try:
            courseId = request.data.get('courseId')  # 班级id
            studentId = request.data.get('studentId')  # 学生id
            result = request.data.get('score')  # 分数
            when = request.data.get('when')  # 学期
            courses = m_model.StudentCourseMemberShip.objects.filter(student_id=studentId, course_id=courseId,
                                                                     when=when)
            if courses:
                courses.update(result=result)
            else:
                m_model.StudentCourseMemberShip.objects.create(
                    student_id=studentId, course_id=courseId, when=when, result=result
                )
            # courses.update(result=result)  # 更新数据
            # courses.save()
            return Response({
                'code': 200,
                'data': '',
                'message': '新增学生课程成绩数据成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'message': '新增学生课程成绩数据失败！'
            })

    def put(self, request):
        """
        修改成绩信息
        """
        try:
            courseId = request.data.get('dbID')  # 记录id
            result = request.data.get('value')  # 成绩
            print(courseId, result)
            m_model.StudentCourseMemberShip.objects.filter(id=courseId).update(result=result)
            # studentId = request.data.get('studentId')  # 班级id
            # result = request.data.get('score')  # 班级id
            # when = request.data.get('when')  # 班级id
            # courses = m_model.StudentCourseMemberShip.objects.filter(student_id=studentId, course_id=courseId,  when=when)  # 查询带有学期的数据记录
            # if courses:  # 如果没有，则查询一下有没有不带学期的记录
            #     courses.update(result=result)  # 更新数据
            # else:
            #     courses = m_model.StudentCourseMemberShip.objects.filter(student_id=studentId, course_id=courseId)
            #     courses.update(result=result, when=when)
            # courses.save()
            return Response({
                'code': 0,
                'data': '',
                'msg': '修改学生课程成绩数据成功！'
            })
        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '修改学生课程成绩数据失败！'
            })


def banji_score(request):
    return render(request, 'student_score/banji_score_main.html')


class banji_score_count(views.APIView):
    def get(self, request):
        fsd = {}
        # 分班级统计学生总人数,总分,平均分
        zrs = models.StudentCourseMemberShip.objects.values("course_id").annotate(zrsCount=Count('student_id'),
                                                                                  zfCount=Sum('result'),
                                                                                  avgCount=Avg('result')).values(
            "college__name", "course_id", "course__grade__name", "student__t_class__name", "course__name", "zrsCount",
            "zfCount", "avgCount")

        # 分班级统计学生及格人数
        jgrs = models.StudentCourseMemberShip.objects.filter(result__gte=60).values("course_id").annotate(
            jgrsCount=Count('*')).values("course_id", "jgrsCount")

        f_0_40 = models.StudentCourseMemberShip.objects.filter(Q(result__gte=0) & Q(result__lt=40)).values(
            "course_id").annotate(
            f_1_40Count=Count('*')).values("course_id", "f_1_40Count")
        f_40_59 = models.StudentCourseMemberShip.objects.filter(Q(result__gte=40) & Q(result__lt=60)).values(
            "course_id").annotate(
            f_40_59Count=Count('*')).values("course_id", "f_40_59Count")
        f_60_79 = models.StudentCourseMemberShip.objects.filter(Q(result__gte=60) & Q(result__lt=80)).values(
            "course_id").annotate(
            f_60_79Count=Count('*')).values("course_id", "f_60_79Count")
        f_80_89 = models.StudentCourseMemberShip.objects.filter(Q(result__gte=80) & Q(result__lt=90)).values(
            "course_id").annotate(
            f_80_89Count=Count('*')).values("course_id", "f_80_89Count")
        f_90_100 = models.StudentCourseMemberShip.objects.filter(Q(result__gte=90) & Q(result__lte=100)).values(
            "course_id").annotate(
            f_90_100Count=Count('*')).values("course_id", "f_90_100Count")

        print(f_40_59)

        total = []
        for x in zrs:
            cj_lit = {'course_id': x['course_id'], 'college': x['college__name'], 'grade': x['course__grade__name'],
                      'class': x['student__t_class__name'], 'course': x['course__name'], 'zrs': x['zrsCount'],
                      'zf': x['zfCount'], 'avg': x['avgCount'], 'jgrs': 0, 'jgl': 0, 'f_1_40Count': 0,
                      'f_40_59Count': 0, 'f_60_79Count': 0, 'f_80_89Count': 0, 'f_90_100Count': 0, }
            total.append(cj_lit)
        for itm in jgrs:
            for i in total:
                if i['course_id'] == itm['course_id']:
                    i['jgrs'] = int(itm['jgrsCount'])  # 插入及格人数
                    i['jgl'] = '{:.1%}'.format(int(itm['jgrsCount']) / int(i['zrs']))  # 计算及格率

        for itm in f_0_40:
            for i in total:
                if i['course_id'] == itm['course_id']:
                    i['f_1_40Count'] = int(itm['f_1_40Count'])  # 插入1-40人数
        for itm in f_40_59:
            for i in total:
                if i['course_id'] == itm['course_id']:
                    i['f_40_59Count'] = int(itm['f_40_59Count'])  # 插入1-40人数
        for itm in f_60_79:
            for i in total:
                if i['course_id'] == itm['course_id']:
                    i['f_60_79Count'] = int(itm['f_60_79Count'])  # 插入1-40人数
        for itm in f_80_89:
            for i in total:
                if i['course_id'] == itm['course_id']:
                    i['f_80_89Count'] = int(itm['f_80_89Count'])  # 插入1-40人数
        for itm in f_90_100:
            for i in total:
                if i['course_id'] == itm['course_id']:
                    i['f_90_100Count'] = int(itm['f_90_100Count'])  # 插入1-40人数

        return Response({
            'code': 0,
            'data': total,
            'msg': '班级成绩数据获取成功！'
        })


def stu_score_manage(request):
    return render(request, 'student_score/stu_score_main.html')


class stu_ScoreManage(views.APIView):
    """
    学生成绩列表
    """

    def get(self, request):
        """
        获取成绩列表
        """
        try:
            # 当前页
            page = request.GET.get('page', 1)
            # 每页条数
            size = request.GET.get('limit', 1)
            # 开始位置
            data_start = (int(page) - 1) * int(size)
            # 结束位置
            data_end = (int(page)) * int(size)
            data = []
            zf = m_model.StudentCourseMemberShip.objects.all()[data_start:data_end].values('course__grade__name',
                                                                                           'student__t_class__name',
                                                                                           'student__name',
                                                                                           'course__name', 'result')
            list = pd.DataFrame(zf)
            ls = pd.pivot_table(list, values='result', index=['course__name'],
                                columns=['course__grade__name', 'student__t_class__name', 'student__name'],
                                aggfunc=np.sum, fill_value=0)  # 数据透视表 # 通过求和来聚合值 可以使用fill_value参数填充缺失的值
            count = m_model.Student.objects.count()

            l = ls.to_dict()  # 转换成字典格式'course__grade__name','student__t_class__name',
            i = 0
            for j in l:  # 转换成layui数据格式
                data.append(l[j])
                data[i]['grade'] = j[0]
                data[i]['class'] = j[1]
                data[i]['name'] = j[2]
                i += 1

            # 数据分页
            # pg = Pagination()
            # pager_roles = pg.paginate_queryset(queryset=data, request=request, view=self)
            # datas = ser.ScoreSerializer(instance=pager_roles, many=True)
            #  用get_paginated_response方法，前端拿到的返回值中有数据总条数，上一页的链接，下一页的链接。

            # return Response({'code': 0, 'data': data})
            # return pg.get_paginated_response(pager_roles)
            return Response({
                'code': 0,
                'data': data,
                'count': count,
                'msg': '获取学生课程成绩数据成功！'
            })


        except Exception as e:
            return Response({
                'code': 500,
                'data': f'{e}',
                'msg': '获取学生课程成绩数据失败！'
            })


def score_query(request):
    data_list = []
    page = request.POST.get('page', 1)
    limit = request.POST.get('limit', 10)
    post_data_str = request.POST.get('Params', None)
    print(post_data_str)

    if post_data_str is None:
        return res_josn_data.table_api(data=data_list, count=0)
    else:
        post_data = json.loads(post_data_str)
        print(post_data)
        college = post_data['method']
        grade = post_data['URL']
        course = post_data['desc']

        t_class = post_data['t_class']

        filters = {}  # 查询参数构造
        # model或数据库对应字段
        orm_field = ['__gt', '__gte', '__lt', '__lte', '__exact', '__iexact', '__contains', '__name',
                     '__startswith', '__istartswith', '__endswith', '__iendswith', '__range', '__isnull', '__in']
        filed_dict = {0: 'college', 1: 'course'}
        param_list = [college, course, ]

        for i in range(len(param_list)):
            if param_list[i] not in (None, ''):
                db_field = filed_dict[i] + orm_field[7]
                filters[db_field] = param_list[i]
        if grade:
            filters['college__grade__name'] = grade
        if t_class:
            filters['student__t_class__name'] = t_class
        print('filters:', filters)
        user_obj = m_model.StudentCourseMemberShip.objects.filter(**filters).order_by('-id').values('id',
                                                                                                    'college__name',
                                                                                                    'college__grade__name',
                                                                                                    'student__t_class__name',
                                                                                                    'student__name',
                                                                                                    'course__name',
                                                                                                    'result')
        page_data = Paginator(user_obj, limit).page(page)

        # 序号
        count = (int(page) - 1) * int(limit)
        print(user_obj)
        for item in page_data:
            count += 1
            # dis_time = str(item.create_time)[:19]

            item_data = {
                "id": count,
                "uid": item['id'],
                "college": item['college__name'],
                "grade": item['college__grade__name'],
                "class": item['student__t_class__name'],
                "name": item['student__name'],
                "course": item['course__name'],
                "result": item['result'],

            }
            data_list.append(item_data)

        return res_josn_data.table_api(count=len(user_obj), data=data_list)


def class_query(request):
    college_name = request.GET.get('college_id')  # 获取新增学生页面，年级下拉菜单改变时过来的gradeid
    grade_name = request.session.get("position")
    print(college_name,grade_name)
    t_class = m_model.Class.objects.filter(Q(grade__college__name=college_name) & Q(grade__name=grade_name)).values('name')
    print(t_class)
    if t_class:
        datas = ser.ClassSerializer(t_class, many=True)
        print(t_class)
        return res_josn_data.table_api(data=datas.data)
    else:
        return res_josn_data.table_api(data='')